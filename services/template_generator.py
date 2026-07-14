from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

def generate_transaction_template():

    wb = Workbook()

    ws = wb.active
    ws.title = "Transactions"

    headers = [
        "#",
        "Date",
        "Description",
        "Account Type",
        "Account Name",
        "Amount",
        "Payment Method",
        "Invoice No."
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    lists = wb.create_sheet("Lists")

    account_types = [
        "Asset",
        "Liability",
        "Equity",
        "Revenue",
        "Expense"
    ]

    payment_methods = [
        "Cash",
        "GCash",
        "Maya"
    ]

    # These are ALL account names.
    # Validation happens during upload.
    account_names = [
        "Cash",
        "Kitchen Equipment",
        "Inventory",
        "Accounts Payable",
        "Loans Payable",
        "Lease Liability",
        "Owner's Equity",
        "Retained Earnings",
        "Food Sales",
        "Beverage and Snack Sales",
        "Cost of Goods Sold (COGS)",
        "Canteen Rent Expense",
        "Utilities Expense"
    ]

    # Write lists
    for i, item in enumerate(account_types, start=1):
        lists[f"A{i}"] = item

    for i, item in enumerate(account_names, start=1):
        lists[f"B{i}"] = item

    for i, item in enumerate(payment_methods, start=1):
        lists[f"C{i}"] = item

    # Dropdowns
    account_type_validation = DataValidation(
        type="list",
        formula1=f"=Lists!$A$1:$A${len(account_types)}"
    )

    account_name_validation = DataValidation(
        type="list",
        formula1=f"=Lists!$B$1:$B${len(account_names)}"
    )

    payment_validation = DataValidation(
        type="list",
        formula1=f"=Lists!$C$1:$C${len(payment_methods)}"
    )

    ws.add_data_validation(account_type_validation)
    ws.add_data_validation(account_name_validation)
    ws.add_data_validation(payment_validation)

    for row in range(2, 1001):

        account_type_validation.add(ws[f"D{row}"])

        account_name_validation.add(ws[f"E{row}"])

        payment_validation.add(ws[f"G{row}"])

    lists.sheet_state = "hidden"

    # Save workbook to the project folder
    output_path = Path.home() / "Downloads" / "Transaction_Template.xlsx"

    wb.save(output_path)

    print(f"Template saved to: {output_path}")

    return output_path

if __name__ == "__main__":
    generate_transaction_template()